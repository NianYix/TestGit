from faker import Faker
import PySimpleGUI as sg
import sys
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl  import load_workbook
# from openpyxl.styles import Font, Alignment
import os


def get_allName(path,language,col,langText,count):
#     wb = Workbook() 
    filename = r'随机名字.xlsx'
    filepath = path + '/' + filename
    try:
        wb2 = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        wb2 = Workbook() 
    ws = wb2.active
    fake = Faker(locale=language)  
    ws.cell(row=1, column=col, value=langText)
    for i in range(count):
        row=i+2
        ws.cell(row=row, column=col, value=fake.name())
    wb2.save(filepath)
#     fake.user_name()



def del_win():
    # 构建GUI
    layout_login = [[sg.Text('账号：'),sg.Input(key='account_id')],
                [sg.Text('密码：'),sg.Input(password_char='*', key='password')],
                [sg.Button('      确定      '), sg.Button('      关闭      ')]
                ]
    w = sg.Window('请输入账号密码', layout=layout_login)
    while True:
        event, values = w.read()
        if event in (None, '      关闭      '):
            sys.exit('程序关闭')
        # 判断账号密码
        elif values['account_id'] == 'admin' and values['password'] == 'admin':
            break
        else:
            sg.popup('账号密码不正确')

    w.close()
    #[[sg.Text('选择课程表原始文件')], [sg.Input(key='fileName',enable_events=True), sg.FileBrowse('打开',file_types=(("Text Files", "*.xls*"),))],
          
    layout = [[sg.Text('选择解析结果存储地址')], [sg.Input(key='path',enable_events=True), sg.FolderBrowse('浏览')],
              [sg.Text('名字数量')], [sg.Input(key='count',enable_events=True)],
          [sg.Text('程序操作记录:',justification='center')],
          [sg.Output(size=(50, 8))],
          [sg.Button('   开始处理   '), sg.Button('      关闭      ')]]

    window = sg.Window('解析工具', layout)
    while True:
        event, values = window.read()
        if event in (None, '      关闭      '):
            break
        if event == '   开始处理   ':
            #fileName = values['fileName']
            path = values['path']
            count = int(values['count'])
            if os.path.exists(path):
#             if os.path.exists(fileName) and os.path.exists(path):
#                 data = get_data(fileName)
#                 get_class(data,path)
                get_allName(path,'zh_CN',1,'简中',count)
                get_allName(path,'zh_TW',2,'繁中',count)
                get_allName(path,'en_US',3,'英语',count)
                get_allName(path,'de_DE',4,'德语',count)
                get_allName(path,'fr_FR',5,'法语',count)
                get_allName(path,'ja_JP',6,'日语',count)
                get_allName(path,'ko_KR',7,'韩语',count)
                print('随机名字表已经保存完毕....')
#                 get_teacher(data,path)
#                 print('按教师课程表已经保存完毕..')
            else:
                sg.popup('请选择正确的待处理文件及保存路径')
    window.close()
    
if __name__ == "__main__":
    del_win()
