import PySimpleGUI as sg
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os

import warnings
warnings.filterwarnings('ignore')
# 读取课程表原始数据
def get_data(fileName):
    df = pd.read_excel(fileName, header=[1,2,3])
    data = df.melt(id_vars=df.columns[:1].to_list())
    data.columns = ['班级','星期','上下午','第N节课','课程_老师']
    # data['课程_老师'] = data['课程_老师'].apply(lambda s:re.sub('上|必须|走.*','',s))
    data['课程'] = data['课程_老师'].apply(lambda s: s if '上' in s else s[:2])
    data['老师'] = data['课程_老师'].apply(lambda s: '' if '上' in s else s[2:]).str.strip().replace('','无')
    data['星期'] = data['星期'].str.strip()
    # data['课程路径'] = data['班级'].astype('str') + '班\n' + data['课程']
    
    return data

# 解析班级课表并保存
def get_class(data,path):
    wb = Workbook()
    ws = wb.active
    # 按班级
    for i in data['班级'].unique():
        result = data.query(f'班级=={i}').pivot(index=['第N节课'],columns='星期',values='课程')
        result = result[['星期一', '星期二', '星期三', '星期四', '星期五']]
        result.index = result.index.map(lambda x: f'第{x}节')
        result.index.name = ''
        result.reset_index(inplace=True)
        result = result.append(pd.Series({'':'第8节', '星期一':'班/安'}),ignore_index=True)
        # 文件名
        filename = r'按班级课程表.xlsx'
        name = f'{i}班'
        filepath = path + '/' + filename
        # 构建excel表格内容
        title = f'课程表  {name}'
        value_list = [[title]]
        
        value_list.append(result.columns.to_list())
        temp = result.values.tolist()
        temp.insert(4,[])
        value_list.extend(temp)
        
        # 写入数据
        for row in value_list:
            ws.append(row)
        # 合并单元格
        ws.merge_cells(f'A{(i-1)*11+1}:F{(i-1)*11+1}')
        cell = ws[f'A{(i-1)*11+1}']
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        print(f'{name} 的课表已生成')
    wb.save(filepath)

# 解析老师课表并保存
def get_teacher(data,path):
    wb = Workbook()
    ws = wb.active
    # 按老师
    for n,i in enumerate(data['老师'].unique()):
        try:
            result = data.query(f'老师=="{i}"').pivot(index=['第N节课'],columns='星期',values='班级')
        except:
            continue
        # 补全星期
        weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五']
        for week in weekdays:
            if week not in result.columns:
                result[week] = ''
        result = result[weekdays]
        # 补全课时
        indexs = range(1,8)

        for index in indexs:
            if index not in result.index:
                result.loc[index,:] = ''
        result.sort_index(inplace=True)
        result.index.name = ''
        result.loc[8,:]=''
        result.reset_index(inplace=True)
        # 文件名称
        filename = r'按老师课程表.xlsx'
        name = f'{i}'
        filepath = path + '/' + filename
        # 构建excel表格内容
        value_list = []
        title1 = '任课教师'
        title2 = i
        title3 = '学科'
        title4 = data.query(f'老师=="{i}"')['课程'].unique()[0]
        title = [[title1,'',title2,'',title3,title4],[]]
        value_list.extend(title)
        value_list.append(result.columns.to_list())
        temp = result.values.tolist()
        temp.insert(4,[])
        value_list.extend(temp)
        
        # 写入数据
        for row in range((n//4)*14+1, (n//4+1)*14+1):        
            for col in range(n%4*7+1, n%4*7+7):
                if row == (n//4)*14+1:
                    a = ws.cell(row=row, column=col)
                    a.font = Font(bold = True)
                else:
                    a = ws.cell(row=row, column=col)
                    a.alignment = Alignment(horizontal="center")
                try:
                    ws.cell(row=row, column=col, value=value_list[row-14*(n//4)-1][col-7*(n%4)-1])
                except:
                    ws.cell(row=row, column=col, value='')
        print(f'{name}老师 的课表已生成')     
    wb.save(filepath)  
 
def del_win():
    # 构建GUI
    layout_login = [[sg.Text('账号：'),sg.Input(key='account_id')],
                [sg.Text('密码：'),sg.Input(password_char='*', key='password')],
                [sg.Button('      确定      '), sg.Button('      关闭      ')]
                ]
    w = sg.Window('请输入账号密码', layout=layout_login)
    while True:
        event, values = w.read()
        if event in (None, '      关闭      '):
            sys.exit('程序关闭')
        # 判断账号密码
        elif values['account_id'] == 'admin' and values['password'] == 'admin':
            break
        else:
            sg.popup('账号密码不正确')

    w.close()
    layout = [[sg.Text('选择课程表原始文件')], [sg.Input(key='fileName',enable_events=True), sg.FileBrowse('打开',file_types=(("Text Files", "*.xls*"),))],
          [sg.Text('选择解析结果存储地址')], [sg.Input(key='path',enable_events=True), sg.FolderBrowse('浏览')],
          [sg.Text('程序操作记录:',justification='center')],
          [sg.Output(size=(50, 8))],
          [sg.Button('   开始处理   '), sg.Button('      关闭      ')]]

    window = sg.Window('课程表解析工具', layout)
    while True:
        event, values = window.read()
        if event in (None, '      关闭      '):
            break
        if event == '   开始处理   ':
            fileName = values['fileName']
            path = values['path']
            if os.path.exists(fileName) and os.path.exists(path):
                data = get_data(fileName)
                get_class(data,path)
                print('按班级课程表已经保存完毕..')
                get_teacher(data,path)
                print('按教师课程表已经保存完毕..')
            else:
                sg.popup('请选择正确的待处理文件及保存路径')
    window.close()
    
 
if __name__ == "__main__":
    del_win()
